import streamlit as st
import pandas as pd
import numpy as np
import io
import matplotlib.pyplot as plt
import networkx as nx
import plotly.express as px
import plotly.graph_objects as go
import random
from datetime import datetime, timedelta, date
import os
import re
import sqlite3
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

# Configurazione pagina
st.set_page_config(
    page_title="Hotel Upgrade Advisor Pro",
    page_icon="🏨",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Funzioni di utilità
def get_data_path(filename):
    """Restituisce il percorso corretto per il file dati"""
    if os.path.exists(filename):
        return filename
    elif os.path.exists(os.path.join("data", filename)):
        return os.path.join("data", filename)
    else:
        os.makedirs("data", exist_ok=True)
        return os.path.join("data", filename)

def get_connection():
    """Crea una connessione al database SQLite"""
    db_path = get_data_path("hotel_data.db")
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    return conn

def ensure_date(date_object):
    """Converte l'oggetto in date se è datetime, se è già date lo lascia invariato"""
    if hasattr(date_object, 'date'):
        return date_object.date()
    return date_object

def init_database():
    """Inizializza il database se non esiste"""
    conn = get_connection()
    cursor = conn.cursor()
    
    # Tabella per dati di pickup
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS pickup_data (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        hotel TEXT NOT NULL,
        stay_date TEXT NOT NULL,
        roomnights INTEGER,
        vs_1day INTEGER,
        vs_2day INTEGER,
        vs_3day INTEGER,
        vs_7day INTEGER,
        vs_7day_spit INTEGER,
        insertion_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )
    ''')
    
    # Tabella per dati di produzione
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS production_data (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        hotel TEXT NOT NULL,
        date TEXT NOT NULL,
        weekday TEXT,
        room_nights INTEGER,
        occupancy REAL,
        adr REAL,
        room_revenue REAL,
        vs_spit_rn INTEGER,
        vs_spit_occ REAL,
        vs_spit_adr REAL,
        vs_ap_rn INTEGER,
        vs_ap_occ REAL,
        vs_ap_adr REAL,
        insertion_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )
    ''')
    
    # Tabella per configurazione hotel
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS hotels (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        hotel_name TEXT UNIQUE NOT NULL,
        total_rooms INTEGER DEFAULT 0,
        address TEXT,
        city TEXT,
        stars INTEGER DEFAULT 4,
        seasonal BOOLEAN DEFAULT 0,
        creation_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )
    ''')
    
    # Tabella per tipologie camere
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS room_types (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        hotel_id INTEGER NOT NULL,
        room_type_name TEXT NOT NULL,
        rooms_in_type INTEGER DEFAULT 0,
        is_entry_level BOOLEAN DEFAULT 0,
        adr REAL DEFAULT 0,
        min_margin REAL DEFAULT 0.6,
        upgrade_threshold REAL DEFAULT 100,
        upgrade_target_type TEXT,
        creation_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (hotel_id) REFERENCES hotels (id),
        UNIQUE (hotel_id, room_type_name)
    )
    ''')
    
    # Tabella per camere Out of Order
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS rooms_ooo (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        hotel_id INTEGER NOT NULL,
        room_type_id INTEGER NOT NULL,
        date_from TEXT NOT NULL,
        date_to TEXT NOT NULL,
        rooms_count INTEGER DEFAULT 1,
        reason TEXT,
        creation_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (hotel_id) REFERENCES hotels (id),
        FOREIGN KEY (room_type_id) REFERENCES room_types (id)
    )
    ''')
    
    # Tabella per parametri hotel - da mantenere per compatibilità
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS hotel_settings (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        hotel TEXT UNIQUE NOT NULL,
        adr REAL DEFAULT 381.93,
        min_margin REAL DEFAULT 0.6,
        upgrade_threshold REAL DEFAULT 100,
        days INTEGER DEFAULT 30,
        is_entry_level BOOLEAN DEFAULT 0,
        room_type_name TEXT,
        rooms_in_type INTEGER DEFAULT 0,
        total_rooms INTEGER DEFAULT 0,
        insertion_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        last_update TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )
    ''')
    
    conn.commit()
    conn.close()

# Funzioni per la gestione degli hotel
def get_all_hotels():
    """Recupera tutti gli hotel dal database"""
    conn = get_connection()
    cursor = conn.cursor()
    
    try:
        cursor.execute("SELECT * FROM hotels ORDER BY hotel_name")
        results = cursor.fetchall()
        return [dict(r) for r in results]
    except Exception as e:
        print(f"Errore nel recupero degli hotel: {e}")
        return []
    finally:
        conn.close()

def get_hotel_by_name(hotel_name):
    """Recupera un hotel dal database tramite nome"""
    conn = get_connection()
    cursor = conn.cursor()
    
    try:
        cursor.execute("SELECT * FROM hotels WHERE hotel_name = ?", (hotel_name,))
        result = cursor.fetchone()
        return dict(result) if result else None
    except Exception as e:
        print(f"Errore nel recupero dell'hotel: {e}")
        return None
    finally:
        conn.close()

def add_hotel(hotel_data):
    """Aggiunge un nuovo hotel al database"""
    conn = get_connection()
    cursor = conn.cursor()
    
    try:
        cursor.execute("""
        INSERT INTO hotels 
        (hotel_name, total_rooms, address, city, stars, seasonal) 
        VALUES (?, ?, ?, ?, ?, ?)
        """, (
            hotel_data.get('hotel_name', ''),
            hotel_data.get('total_rooms', 0),
            hotel_data.get('address', ''),
            hotel_data.get('city', ''),
            hotel_data.get('stars', 4),
            hotel_data.get('seasonal', False)
        ))
        conn.commit()
        
        # Restituisci l'ID dell'hotel appena inserito
        cursor.execute("SELECT last_insert_rowid()")
        result = cursor.fetchone()
        return result[0] if result else None
    except Exception as e:
        conn.rollback()
        print(f"Errore nell'inserimento dell'hotel: {e}")
        return None
    finally:
        conn.close()

def update_hotel(hotel_id, hotel_data):
    """Aggiorna i dati di un hotel esistente"""
    conn = get_connection()
    cursor = conn.cursor()
    
    try:
        cursor.execute("""
        UPDATE hotels SET
        hotel_name = ?,
        total_rooms = ?,
        address = ?,
        city = ?,
        stars = ?,
        seasonal = ?
        WHERE id = ?
        """, (
            hotel_data.get('hotel_name'),
            hotel_data.get('total_rooms'),
            hotel_data.get('address'),
            hotel_data.get('city'),
            hotel_data.get('stars'),
            hotel_data.get('seasonal'),
            hotel_id
        ))
        conn.commit()
        return True
    except Exception as e:
        conn.rollback()
        print(f"Errore nell'aggiornamento dell'hotel: {e}")
        return False
    finally:
        conn.close()

def delete_hotel(hotel_id):
    """Elimina un hotel dal database"""
    conn = get_connection()
    cursor = conn.cursor()
    
    try:
        # Elimina prima tutte le tipologie associate
        cursor.execute("DELETE FROM room_types WHERE hotel_id = ?", (hotel_id,))
        # Elimina l'hotel
        cursor.execute("DELETE FROM hotels WHERE id = ?", (hotel_id,))
        conn.commit()
        return True
    except Exception as e:
        conn.rollback()
        print(f"Errore nell'eliminazione dell'hotel: {e}")
        return False
    finally:
        conn.close()

# Funzioni per la gestione delle tipologie di camera
def get_room_types_by_hotel(hotel_id):
    """Recupera tutte le tipologie di camera di un hotel"""
    conn = get_connection()
    cursor = conn.cursor()
    
    try:
        cursor.execute("""
        SELECT * FROM room_types 
        WHERE hotel_id = ?
        ORDER BY adr
        """, (hotel_id,))
        results = cursor.fetchall()
        return [dict(r) for r in results]
    except Exception as e:
        print(f"Errore nel recupero delle tipologie: {e}")
        return []
    finally:
        conn.close()

def get_room_type_by_name(hotel_id, room_type_name):
    """Recupera una tipologia di camera tramite nome"""
    conn = get_connection()
    cursor = conn.cursor()
    
    try:
        cursor.execute("""
        SELECT * FROM room_types 
        WHERE hotel_id = ? AND room_type_name = ?
        """, (hotel_id, room_type_name))
        result = cursor.fetchone()
        return dict(result) if result else None
    except Exception as e:
        print(f"Errore nel recupero della tipologia: {e}")
        return None
    finally:
        conn.close()

def add_room_type(hotel_id, room_type_data):
    """Aggiunge una nuova tipologia di camera"""
    conn = get_connection()
    cursor = conn.cursor()
    
    try:
        cursor.execute("""
        INSERT INTO room_types 
        (hotel_id, room_type_name, rooms_in_type, is_entry_level, adr, min_margin, 
         upgrade_threshold, upgrade_target_type)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            hotel_id,
            room_type_data.get('room_type_name', ''),
            room_type_data.get('rooms_in_type', 0),
            room_type_data.get('is_entry_level', False),
            room_type_data.get('adr', 0),
            room_type_data.get('min_margin', 0.6),
            room_type_data.get('upgrade_threshold', 100),
            room_type_data.get('upgrade_target_type')
        ))
        conn.commit()
        
        # Restituisci l'ID della tipologia appena inserita
        cursor.execute("SELECT last_insert_rowid()")
        result = cursor.fetchone()
        return result[0] if result else None
    except Exception as e:
        conn.rollback()
        print(f"Errore nell'inserimento della tipologia: {e}")
        return None
    finally:
        conn.close()

def update_room_type(room_type_id, room_type_data):
    """Aggiorna i dati di una tipologia esistente"""
    conn = get_connection()
    cursor = conn.cursor()
    
    try:
        cursor.execute("""
        UPDATE room_types SET
        room_type_name = ?,
        rooms_in_type = ?,
        is_entry_level = ?,
        adr = ?,
        min_margin = ?,
        upgrade_threshold = ?,
        upgrade_target_type = ?
        WHERE id = ?
        """, (
            room_type_data.get('room_type_name'),
            room_type_data.get('rooms_in_type'),
            room_type_data.get('is_entry_level'),
            room_type_data.get('adr'),
            room_type_data.get('min_margin'),
            room_type_data.get('upgrade_threshold'),
            room_type_data.get('upgrade_target_type'),
            room_type_id
        ))
        conn.commit()
        return True
    except Exception as e:
        conn.rollback()
        print(f"Errore nell'aggiornamento della tipologia: {e}")
        return False
    finally:
        conn.close()

# Funzioni per la gestione delle camere Out of Order (OOO)
def get_active_ooo_rooms(hotel_id):
    """Recupera le camere fuori servizio attive"""
    conn = get_connection()
    cursor = conn.cursor()
    
    try:
        cursor.execute("""
        SELECT o.*, r.room_type_name 
        FROM rooms_ooo o
        JOIN room_types r ON o.room_type_id = r.id
        WHERE o.hotel_id = ? AND date_to >= date('now')
        ORDER BY date_from
        """, (hotel_id,))
        results = cursor.fetchall()
        return [dict(r) for r in results]
    except Exception as e:
        print(f"Errore nel recupero delle camere OOO: {e}")
        return []
    finally:
        conn.close()

def add_ooo_rooms(hotel_id, room_type_id, ooo_data):
    """Aggiunge camere fuori servizio"""
    conn = get_connection()
    cursor = conn.cursor()
    
    try:
        cursor.execute("""
        INSERT INTO rooms_ooo
        (hotel_id, room_type_id, date_from, date_to, rooms_count, reason)
        VALUES (?, ?, ?, ?, ?, ?)
        """, (
            hotel_id,
            room_type_id,
            ooo_data.get('date_from'),
            ooo_data.get('date_to'),
            ooo_data.get('rooms_count', 1),
            ooo_data.get('reason', '')
        ))
        conn.commit()
        return True
    except Exception as e:
        conn.rollback()
        print(f"Errore nell'inserimento delle camere OOO: {e}")
        return False
    finally:
        conn.close()

def end_ooo(ooo_id):
    """Termina un periodo di OOO anticipatamente"""
    conn = get_connection()
    cursor = conn.cursor()
    
    try:
        # Imposta la data di fine a ieri
        yesterday = (datetime.now() - timedelta(days=1)).strftime('%d/%m/%Y')
        cursor.execute("""
        UPDATE rooms_ooo 
        SET date_to = ?
        WHERE id = ? AND date_to > date('now')
        """, (yesterday, ooo_id))
        conn.commit()
        return True
    except Exception as e:
        conn.rollback()
        print(f"Errore nella terminazione dell'OOO: {e}")
        return False
    finally:
        conn.close()

def get_effective_capacity(hotel_id, room_type_id, date):
    """
    Calcola la capacità effettiva di una tipologia per una data specifica,
    considerando le camere fuori servizio.
    """
    try:
        conn = get_connection()
        cursor = conn.cursor()
        
        # Ottieni capacità base
        cursor.execute("""
        SELECT rooms_in_type FROM room_types 
        WHERE id = ? AND hotel_id = ?
        """, (room_type_id, hotel_id))
        
        result = cursor.fetchone()
        if not result:
            return 0
        
        base_capacity = result['rooms_in_type']
        
        # Trova camere OOO per questa data
        date_str = date.strftime('%d/%m/%Y') if isinstance(date, datetime) else date
        
        cursor.execute("""
        SELECT SUM(rooms_count) as ooo_rooms FROM rooms_ooo
        WHERE hotel_id = ? AND room_type_id = ?
        AND date_from <= ? AND date_to >= ?
        """, (hotel_id, room_type_id, date_str, date_str))
        
        ooo_result = cursor.fetchone()
        ooo_count = ooo_result['ooo_rooms'] if ooo_result and ooo_result['ooo_rooms'] else 0
        
        # Capacità effettiva
        effective_capacity = max(0, base_capacity - ooo_count)
        
        return effective_capacity
    
    except Exception as e:
        print(f"Errore nel calcolo della capacità effettiva: {e}")
        return 0
    finally:
        conn.close()

# Funzioni per l'analisi dei dati
def analyze_vs_spit_trend(hotel_name, days=60):
    """
    Analizza il trend vs SPIT nelle ultime settimane per un hotel.
    
    Args:
        hotel_name: Nome dell'hotel
        days: Numero di giorni da analizzare
    
    Returns:
        Dizionario con statistiche sul trend vs SPIT
    """
    daily_data = get_latest_excel_data(hotel=hotel_name, section="Produzione Giornaliera")
    if not daily_data or "daily_data" not in daily_data:
        return {"error": "Dati non disponibili"}
    
    df = pd.DataFrame(daily_data["daily_data"])
    df["date_obj"] = pd.to_datetime(df["date"], format="%d/%m/%Y", errors="coerce")
    
    if "vs_spit_rn" not in df.columns or "room_nights" not in df.columns:
        return {"error": "Dati vs SPIT non disponibili"}
    
    # Converti a numerico
    df["vs_spit_rn"] = pd.to_numeric(df["vs_spit_rn"], errors="coerce")
    df["room_nights"] = pd.to_numeric(df["room_nights"], errors="coerce")
    
    today = datetime.now().date()
    cutoff_date = today - timedelta(days=days)
    
    recent_data = df[df["date_obj"].dt.date >= cutoff_date]
    
    # Totali per il periodo
    total_rn = recent_data["room_nights"].sum()
    total_vs_spit = recent_data["vs_spit_rn"].sum()
    
    # Calcola i roomnights dell'anno scorso
    spit_rn = total_rn - total_vs_spit
    
    # Calcola variazione percentuale
    vs_spit_pct = (total_vs_spit / spit_rn * 100) if spit_rn > 0 else 0
    
    return {
        "total_rn": total_rn,
        "total_vs_spit": total_vs_spit,
        "spit_rn": spit_rn,
        "vs_spit_pct": vs_spit_pct,
        "period_days": days
    }

# Funzioni per l'estrazione e l'elaborazione dei dati PowerBI
def excel_column_to_index(column_letter):
    """Converte lettera colonna Excel in indice numerico (base 0)."""
    result = 0
    for char in column_letter:
        result = result * 26 + (ord(char.upper()) - ord('A') + 1)
    return result - 1

def detect_report_type(df):
    """
    Identifica automaticamente il tipo di report basandosi sulla struttura delle colonne.
    """
    # Converti i nomi delle colonne in lowercase per confronto case-insensitive
    cols_lower = [str(col).lower() for col in df.columns]
    
    # Check if it's a pickup report
    pickup_cols = ["vs 1gg", "vs 2gg", "vs 3gg", "vs 7gg", "vs 7gg SPIT"]
    pickup_cols_found = sum(1 for col in pickup_cols if any(col.lower() in str(c).lower() for c in df.columns))
    if ("soggiorno" in cols_lower or "data" in cols_lower) and pickup_cols_found >= 2:
        return "Pickup", pickup_cols_found
    
    # Schema colonne produzione portafoglio mensile
    if any("mese" in col for col in cols_lower) and any("occ.%" in col for col in cols_lower or "occupazione" in col for col in cols_lower):
        confidence = sum(1 for pattern in ["room revenue", "adr", "revpar", "mese"] 
                         if any(pattern in col for col in cols_lower))
        if confidence >= 3:
            return "Produzione Portafoglio", confidence
    
    # Schema colonne produzione giornaliera
    if any("data" in col for col in cols_lower) and any("giorno" in col for col in cols_lower):
        confidence = sum(1 for pattern in ["% occ", "room nights", "adr", "data"] 
                         if any(pattern in col for col in cols_lower))
        if confidence >= 3:
            return "Produzione Giornaliera", confidence
    
    # Schema colonne produzione segmenti
    if any("segment" in col for col in cols_lower) and any(("perc" in col or "%" in col) for col in cols_lower):
        return "Prod. per Segmento", 3
    
    # Non è stato possibile identificare il tipo
    return None, 0

def extract_room_type_from_powerbi(file, hotel_name="Cala Cuncheddi"):
    """
    Estrae la tipologia di camera specifica dal report PowerBI del Cala Cuncheddi.
    """
    try:
        # Legge le prime righe del file
        df = pd.read_excel(file, nrows=20)
        
        # Lista completa delle tipologie del Cala Cuncheddi
        cala_cuncheddi_room_types = [
            "Classic Garden",
            "Classic Sea View",
            "Family",
            "Superior Sea View",
            "Executive",
            "Deluxe",
            "Junior Suite Pool",
            "Suite"
        ]
        
        # Cerca la tipologia di camera nei filtri
        for idx, row in df.iterrows():
            row_str = ' '.join([str(val) for val in row.values if not pd.isna(val)])
            
            if "desc. tipo camera" in row_str.lower() or "tipo camera venduta" in row_str.lower():
                # Cerca le tipologie esatte
                for room_type in cala_cuncheddi_room_types:
                    if room_type.lower() in row_str.lower():
                        return {
                            'hotel_name': hotel_name,
                            'room_type': room_type
                        }
                
                # Se non trova una corrispondenza esatta, cerca parti di tipologie
                for segment in row_str.lower().split():
                    for room_type in cala_cuncheddi_room_types:
                        if (segment in room_type.lower() and
                            ("classic" in segment or "superior" in segment or 
                             "suite" in segment or "executive" in segment or
                             "deluxe" in segment or "family" in segment)):
                            return {
                                'hotel_name': hotel_name,
                                'room_type': room_type
                            }
        
        # Ricerca più ampia come backup
        for idx, row in df.iterrows():
            for col in row.values:
                if pd.notna(col) and isinstance(col, str):
                    for room_type in cala_cuncheddi_room_types:
                        if room_type.lower() in col.lower():
                            return {
                                'hotel_name': hotel_name,
                                'room_type': room_type
                            }
        
        return None
    except Exception as e:
        print(f"Errore nell'estrazione della tipologia: {e}")
        return None

def process_pickup_excel(uploaded_file, hotel):
    """Process pickup Excel file and extract relevant data."""
    try:
        df = pd.read_excel(uploaded_file)
        
        # Use the exact column names from your Excel file
        expected_columns = ['Soggiorno', 'Roomnights', 'vs 1gg', 'vs 2gg', 'vs 3gg', 'vs 7gg', 'vs 7gg SPIT']
        
        # Verify all expected columns exist
        missing_columns = [col for col in expected_columns if not any(col.lower() in str(c).lower() for c in df.columns)]
        if missing_columns:
            return None, f"Colonne mancanti: {', '.join(missing_columns)}. Verificare il formato."
        
        # Identify the actual column names in the DataFrame
        soggiorno_col = next((col for col in df.columns if 'soggiorno' in str(col).lower()), None)
        roomnights_col = next((col for col in df.columns if 'roomnights' in str(col).lower()), None)
        vs_1gg_col = next((col for col in df.columns if 'vs 1gg' in str(col).lower()), None)
        vs_2gg_col = next((col for col in df.columns if 'vs 2gg' in str(col).lower()), None)
        vs_3gg_col = next((col for col in df.columns if 'vs 3gg' in str(col).lower()), None)
        vs_7gg_col = next((col for col in df.columns if 'vs 7gg' in str(col).lower() and 'spit' not in str(col).lower()), None)
        vs_7gg_spit_col = next((col for col in df.columns if 'vs 7gg spit' in str(col).lower()), None)
        
        if not all([soggiorno_col, roomnights_col, vs_7gg_col, vs_7gg_spit_col]):
            return None, "Alcune colonne essenziali non sono state trovate. Verificare formato file."
        
        pickup_data = []
        
        for idx, row in df.iterrows():
            # Skip header rows or empty rows
            if pd.isna(row[soggiorno_col]) or (isinstance(row[soggiorno_col], str) and "soggiorno" in row[soggiorno_col].lower()):
                continue
                
            # Format date properly
            stay_date = row[soggiorno_col]
            if isinstance(stay_date, (datetime, date)):
                formatted_date = stay_date.strftime('%d/%m/%Y')
            else:
                formatted_date = str(stay_date)
                
            # Create data entry with proper error handling
            try:
                pickup_entry = {
                    "hotel": hotel,
                    "stay_date": formatted_date,
                    "roomnights": int(float(row[roomnights_col])) if pd.notna(row[roomnights_col]) else 0,
                    "vs_1day": int(float(row[vs_1gg_col])) if vs_1gg_col and pd.notna(row[vs_1gg_col]) else 0,
                    "vs_2day": int(float(row[vs_2gg_col])) if vs_2gg_col and pd.notna(row[vs_2gg_col]) else 0,
                    "vs_3day": int(float(row[vs_3gg_col])) if vs_3gg_col and pd.notna(row[vs_3gg_col]) else 0,
                    "vs_7day": int(float(row[vs_7gg_col])) if pd.notna(row[vs_7gg_col]) else 0,
                    "vs_7day_spit": int(float(row[vs_7gg_spit_col])) if pd.notna(row[vs_7gg_spit_col]) else 0
                }
                pickup_data.append(pickup_entry)
            except Exception as e:
                st.warning(f"Errore riga {idx}: {str(e)}")
                continue
                
        if not pickup_data:
            return None, "Nessun dato valido trovato nel file"
            
        return pickup_data, None
    except Exception as e:
        return None, f"Errore nell'elaborazione: {str(e)}"

def process_daily_production_excel(uploaded_file, hotel):
    """
    Elabora il file Excel di produzione giornaliera caricato.
    """
    try:
        df = pd.read_excel(uploaded_file)
        
        # Mappatura colonne per il report di produzione giornaliera
        expected_structure = {
            'date_col': {'names': ['Data', 'data', 'date'], 'index': 0, 'letter': 'A'},
            'weekday_col': {'names': ['Giorno', 'giorno', 'day'], 'index': 1, 'letter': 'B'},
            'occ_col': {'names': ['% Occ.', '% occ', 'occupancy', 'occupazione'], 'index': 2, 'letter': 'C'},
            'occ_vs_spit_col': {'names': ['vs SPIT', 'vs spit'], 'index': 3, 'letter': 'D'},
            'occ_vs_ap_col': {'names': ['vs AP', 'vs ap'], 'index': 4, 'letter': 'E'},
            'room_nights_col': {'names': ['Room nights', 'room nights', 'roomnights'], 'index': 5, 'letter': 'F'},
            'rn_vs_spit_col': {'names': ['vs SPIT', 'vs spit'], 'index': 6, 'letter': 'G'},
            'rn_vs_ap_col': {'names': ['vs AP', 'vs ap'], 'index': 7, 'letter': 'H'},
            'adr_col': {'names': ['ADR Cam', 'adr cam', 'adr'], 'index': 11, 'letter': 'L'},
            'adr_vs_spit_col': {'names': ['vs SPIT', 'vs spit'], 'index': 12, 'letter': 'M'},
            'adr_vs_ap_col': {'names': ['vs AP', 'vs ap'], 'index': 13, 'letter': 'N'},
            'room_rev_col': {'names': ['Room Revenue', 'room revenue', 'revenue'], 'index': 17, 'letter': 'R'},
            'rev_vs_spit_col': {'names': ['vs SPIT', 'vs spit'], 'index': 18, 'letter': 'S'},
            'rev_vs_ap_col': {'names': ['vs AP', 'vs ap'], 'index': 19, 'letter': 'T'}
        }
        
        col_map = {}
        column_list = list(df.columns)
        num_columns = len(column_list)
        
        # Funzione per trovare colonne per nome o posizione
        def find_column(target_names, context="", letter=None):
            context_lower = context.lower()
            
            # Prova a trovare per lettera della colonna
            if letter:
                try:
                    idx = excel_column_to_index(letter)
                    if idx < num_columns:
                        return column_list[idx]
                except:
                    pass
            
            # Prova a trovare per nome esatto
            for name in target_names:
                if name in column_list:
                    return name
            
            # Prova a trovare per nome case-insensitive
            for col in column_list:
                col_str = str(col).lower()
                for name in target_names:
                    if name.lower() == col_str:
                        return col
            
            # Prova a trovare con contesto
            if context:
                for col in column_list:
                    col_str = str(col).lower()
                    for name in target_names:
                        if name.lower() in col_str and context_lower in col_str:
                            return col
            
            # Prova a trovare per contenuto parziale
            for col in column_list:
                col_str = str(col).lower()
                for name in target_names:
                    if name.lower() in col_str:
                        return col
            
            return None
        
        # Costruisci la mappatura delle colonne
        for key, config in expected_structure.items():
            context = ""
            if "adr" in key:
                context = "adr"
            elif "occ" in key:
                context = "occ"
            elif "room_nights" in key or "rn" in key:
                context = "room nights"
            elif "rev" in key:
                context = "revenue"
            
            letter = config.get('letter')
            index = config['index']
            
            # Prova prima per lettera, poi per indice, poi per nome
            if letter:
                try:
                    idx = excel_column_to_index(letter)
                    if idx < num_columns:
                        col_map[key] = column_list[idx]
                        continue
                except:
                    pass
            
            if index < num_columns:
                col_map[key] = column_list[index]
            else:
                found_col = find_column(config['names'], context)
                if found_col:
                    col_map[key] = found_col
        
        if 'date_col' not in col_map or col_map['date_col'] not in df.columns:
            return None, "Colonna data non trovata. Verificare formato file."
        
        # Pulizia dati: rimuovi righe vuote e intestazioni duplicate
        df = df.dropna(subset=[col_map['date_col']])
        
        # Normalizza percentuali
        if 'occ_col' in col_map and col_map['occ_col'] in df.columns:
            occ_col = col_map['occ_col']
            if df[occ_col].dtype == 'object':
                df[occ_col] = df[occ_col].astype(str).str.replace('%', '').str.replace(',', '.').astype(float)
            if df[occ_col].max() <= 1.0:
                df[occ_col] = df[occ_col] * 100
        
        # Estrai dati giornalieri
        daily_data = []
        
        for idx, row in df.iterrows():
            try:
                # Verifica che sia una riga valida con data
                if pd.isna(row[col_map['date_col']]):
                    continue
                
                date_val = row[col_map['date_col']]
                
                # Salta righe con filtri o intestazioni
                if isinstance(date_val, str) and ("filtri" in date_val.lower() or "data" in date_val.lower()):
                    continue
                
                # Formatta la data
                if isinstance(date_val, (datetime, date)):
                    formatted_date = date_val.strftime('%d/%m/%Y')
                else:
                    formatted_date = str(date_val)
                
                # Estrai i dati richiesti con gestione errori
                daily_entry = {
                    "date": formatted_date,
                    "hotel": hotel
                }
                
                # Giorno della settimana
                if 'weekday_col' in col_map and pd.notna(row[col_map['weekday_col']]):
                    daily_entry["weekday"] = str(row[col_map['weekday_col']])
                
                # Room nights
                if 'room_nights_col' in col_map and pd.notna(row[col_map['room_nights_col']]):
                    rn_val = row[col_map['room_nights_col']]
                    if isinstance(rn_val, str):
                        rn_val = rn_val.replace('.', '').replace(',', '.')
                    daily_entry["room_nights"] = int(float(rn_val))
                
                # Occupazione
                if 'occ_col' in col_map and pd.notna(row[col_map['occ_col']]):
                    occ_val = row[col_map['occ_col']]
                    if isinstance(occ_val, str):
                        occ_val = occ_val.replace('%', '').replace(',', '.')
                    daily_entry["occupancy"] = float(occ_val)
                
                # ADR
                if 'adr_col' in col_map and pd.notna(row[col_map['adr_col']]):
                    adr_val = row[col_map['adr_col']]
                    if isinstance(adr_val, str):
                        adr_val = adr_val.replace('.', '').replace(',', '.')
                    daily_entry["adr"] = float(adr_val)
                
                # Room Revenue
                if 'room_rev_col' in col_map and pd.notna(row[col_map['room_rev_col']]):
                    rev_val = row[col_map['room_rev_col']]
                    if isinstance(rev_val, str):
                        rev_val = rev_val.replace('.', '').replace(',', '.')
                    daily_entry["room_revenue"] = float(rev_val)
                
                # vs SPIT RN
                if 'rn_vs_spit_col' in col_map and pd.notna(row[col_map['rn_vs_spit_col']]):
                    val = row[col_map['rn_vs_spit_col']]
                    if isinstance(val, str):
                        val = val.replace('.', '').replace(',', '.')
                    daily_entry["vs_spit_rn"] = int(float(val))
                
                # vs AP RN
                if 'rn_vs_ap_col' in col_map and pd.notna(row[col_map['rn_vs_ap_col']]):
                    val = row[col_map['rn_vs_ap_col']]
                    if isinstance(val, str):
                        val = val.replace('.', '').replace(',', '.')
                    daily_entry["vs_ap_rn"] = int(float(val))
                
                # vs SPIT Occ
                if 'occ_vs_spit_col' in col_map and pd.notna(row[col_map['occ_vs_spit_col']]):
                    val = row[col_map['occ_vs_spit_col']]
                    if isinstance(val, str):
                        val = val.replace('%', '').replace(',', '.')
                    daily_entry["vs_spit_occ"] = float(val)
                
                # vs AP Occ
                if 'occ_vs_ap_col' in col_map and pd.notna(row[col_map['occ_vs_ap_col']]):
                    val = row[col_map['occ_vs_ap_col']]
                    if isinstance(val, str):
                        val = val.replace('%', '').replace(',', '.')
                    daily_entry["vs_ap_occ"] = float(val)
                
                # vs SPIT ADR
                if 'adr_vs_spit_col' in col_map and pd.notna(row[col_map['adr_vs_spit_col']]):
                    val = row[col_map['adr_vs_spit_col']]
                    if isinstance(val, str):
                        val = val.replace('.', '').replace(',', '.')
                    daily_entry["vs_spit_adr"] = float(val)
                
                # vs AP ADR
                if 'adr_vs_ap_col' in col_map and pd.notna(row[col_map['adr_vs_ap_col']]):
                    val = row[col_map['adr_vs_ap_col']]
                    if isinstance(val, str):
                        val = val.replace('.', '').replace(',', '.')
                    daily_entry["vs_ap_adr"] = float(val)
                
                daily_data.append(daily_entry)
            except Exception as e:
                st.warning(f"Errore processando riga {idx}: {e}")
                continue
        
        if not daily_data:
            return None, "Nessun dato valido trovato nel file di produzione"
        
        return daily_data, None
    except Exception as e:
        return None, f"Errore nell'elaborazione del file di produzione: {str(e)}"

def map_bi_data_to_model(daily_production=None, pickup_data=None):
    """
    Mappa i dati dalla BI al formato richiesto dal modello di upgrade.
    """
    historical_data = {}
    otb_data = {}
    pickup_mapped = {}
    
    # Mapping dati storici dalla produzione giornaliera
    if daily_production:
        # Converti date in oggetti datetime per ordinamento e confronto
        for entry in daily_production:
            try:
                date_str = entry["date"]
                date_parts = date_str.split('/')
                if len(date_parts) == 3:
                    day, month, year = map(int, date_parts)
                    date_obj = datetime(year, month, day)
                    
                    # I dati di produzione rappresentano lo storico
                    historical_data[date_obj] = entry["room_nights"]
            except Exception as e:
                st.warning(f"Errore nel mapping data: {date_str} - {str(e)}")
    
    # Mapping dati pickup
    if pickup_data:
        for entry in pickup_data:
            try:
                date_str = entry["stay_date"]
                date_parts = date_str.split('/')
                if len(date_parts) == 3:
                    day, month, year = map(int, date_parts)
                    date_obj = datetime(year, month, day)
                    
                    # Le room nights correnti dal pickup sono l'OTB
                    otb_data[date_obj] = entry["roomnights"]
                    
                    # Costruisci i dati di pickup
                    pickup_mapped[date_obj] = {
                        "2025": entry["vs_7day"],
                        "2024": entry["vs_7day_spit"]
                    }
            except Exception as e:
                st.warning(f"Errore nel mapping pickup: {date_str} - {str(e)}")
    
    return historical_data, otb_data, pickup_mapped

def save_pickup_data(hotel, pickup_data):
    """Salva i dati di pickup nel database"""
    conn = get_connection()
    cursor = conn.cursor()
    
    saved_count = 0
    try:
        for entry in pickup_data:
            # Verifica se esiste già un record per questa data
            cursor.execute("""
            SELECT id FROM pickup_data 
            WHERE hotel = ? AND stay_date = ?
            """, (hotel, entry['stay_date']))
            
            existing = cursor.fetchone()
            
            if existing:
                # Aggiorna record esistente
                cursor.execute("""
                UPDATE pickup_data 
                SET roomnights = ?, vs_1day = ?, vs_2day = ?, vs_3day = ?, 
                    vs_7day = ?, vs_7day_spit = ?, insertion_date = CURRENT_TIMESTAMP
                WHERE hotel = ? AND stay_date = ?
                """, (
                    entry['roomnights'], entry.get('vs_1day', 0), entry.get('vs_2day', 0), 
                    entry.get('vs_3day', 0), entry.get('vs_7day', 0), entry.get('vs_7day_spit', 0),
                    hotel, entry['stay_date']
                ))
            else:
                # Inserisci nuovo record
                cursor.execute("""
                INSERT INTO pickup_data 
                (hotel, stay_date, roomnights, vs_1day, vs_2day, vs_3day, vs_7day, vs_7day_spit)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    hotel, entry['stay_date'], entry['roomnights'], 
                    entry.get('vs_1day', 0), entry.get('vs_2day', 0), entry.get('vs_3day', 0),
                    entry.get('vs_7day', 0), entry.get('vs_7day_spit', 0)
                ))
            
            saved_count += 1
        
        conn.commit()
        return saved_count
    except Exception as e:
        conn.rollback()
        st.error(f"Errore nel salvataggio dei dati pickup: {e}")
        return 0
    finally:
        conn.close()

def save_production_data(hotel, production_data):
    """Salva i dati di produzione nel database"""
    conn = get_connection()
    cursor = conn.cursor()
    
    saved_count = 0
    try:
        for entry in production_data:
            # Verifica se esiste già un record per questa data
            cursor.execute("""
            SELECT id FROM production_data 
            WHERE hotel = ? AND date = ?
            """, (hotel, entry['date']))
            
            existing = cursor.fetchone()
            
            if existing:
                # Aggiorna record esistente
                cursor.execute("""
                UPDATE production_data 
                SET weekday = ?, room_nights = ?, occupancy = ?, adr = ?, room_revenue = ?,
                    vs_spit_rn = ?, vs_spit_occ = ?, vs_spit_adr = ?,
                    vs_ap_rn = ?, vs_ap_occ = ?, vs_ap_adr = ?,
                    insertion_date = CURRENT_TIMESTAMP
                WHERE hotel = ? AND date = ?
                """, (
                    entry.get('weekday', ''), entry.get('room_nights', 0), 
                    entry.get('occupancy', 0), entry.get('adr', 0), 
                    entry.get('room_revenue', 0),
                    entry.get('vs_spit_rn', 0), entry.get('vs_spit_occ', 0), 
                    entry.get('vs_spit_adr', 0),
                    entry.get('vs_ap_rn', 0), entry.get('vs_ap_occ', 0), 
                    entry.get('vs_ap_adr', 0),
                    hotel, entry['date']
                ))
            else:
                # Inserisci nuovo record
                cursor.execute("""
                INSERT INTO production_data 
                (hotel, date, weekday, room_nights, occupancy, adr, room_revenue,
                vs_spit_rn, vs_spit_occ, vs_spit_adr, vs_ap_rn, vs_ap_occ, vs_ap_adr)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    hotel, entry['date'], entry.get('weekday', ''), 
                    entry.get('room_nights', 0), entry.get('occupancy', 0), 
                    entry.get('adr', 0), entry.get('room_revenue', 0),
                    entry.get('vs_spit_rn', 0), entry.get('vs_spit_occ', 0), 
                    entry.get('vs_spit_adr', 0),
                    entry.get('vs_ap_rn', 0), entry.get('vs_ap_occ', 0), 
                    entry.get('vs_ap_adr', 0)
                ))
            
            saved_count += 1
        
        conn.commit()
        return saved_count
    except Exception as e:
        conn.rollback()
        st.error(f"Errore nel salvataggio dei dati produzione: {e}")
        return 0
    finally:
        conn.close()

def get_latest_excel_data(hotel, section="Produzione Giornaliera"):
    """Recupera i dati più recenti caricati per un hotel e una sezione"""
    if section == "Produzione Giornaliera":
        return {"daily_data": get_production_data(hotel)}
    elif section == "Pickup":
        return {"pickup_data": get_pickup_data(hotel)}
    else:
        return None

def get_pickup_data(hotel, date_from=None, date_to=None):
    """Retrieve pickup data from database with optional date filtering."""
    conn = get_connection()
    c = conn.cursor()
    
    query = "SELECT * FROM pickup_data WHERE hotel = ?"
    params = [hotel]
    
    # Ensure date parameters are properly formatted for comparison
    if date_from:
        # Convert to standard format if needed
        try:
            if isinstance(date_from, (datetime, date)):
                date_from = date_from.strftime('%d/%m/%Y')
            query += " AND stay_date >= ?"
            params.append(date_from)
        except Exception as e:
            print(f"Error formatting date_from parameter: {str(e)}")
    
    if date_to:
        # Convert to standard format if needed
        try:
            if isinstance(date_to, (datetime, date)):
                date_to = date_to.strftime('%d/%m/%Y')
            query += " AND stay_date <= ?"
            params.append(date_to)
        except Exception as e:
            print(f"Error formatting date_to parameter: {str(e)}")
    
    query += " ORDER BY stay_date ASC"
    
    try:
        c.execute(query, params)
        results = c.fetchall()
        
        # Process results for consistency
        processed_results = []
        for row in results:
            row_dict = dict(row)
            
            # Ensure stay_date is properly formatted
            if "stay_date" in row_dict and row_dict["stay_date"]:
                try:
                    # If it's not already a properly formatted string
                    if not isinstance(row_dict["stay_date"], str) or not re.match(r'\d{2}/\d{2}/\d{4}', row_dict["stay_date"]):
                        if isinstance(row_dict["stay_date"], (datetime, date)):
                            row_dict["stay_date"] = row_dict["stay_date"].strftime('%d/%m/%Y')
                except Exception:
                    # Keep as is if conversion fails
                    pass
                    
            processed_results.append(row_dict)
            
        return processed_results
    except Exception as e:
        print(f"Error retrieving pickup data: {str(e)}")
        return []
    finally:
        conn.close()

def get_production_data(hotel, date_from=None, date_to=None):
    """Recupera dati di produzione dal database con filtro opzionale per date"""
    conn = get_connection()
    c = conn.cursor()
    
    query = "SELECT * FROM production_data WHERE hotel = ?"
    params = [hotel]
    
    # Applica filtri per data
    if date_from:
        try:
            if isinstance(date_from, (datetime, date)):
                date_from = date_from.strftime('%d/%m/%Y')
            query += " AND date >= ?"
            params.append(date_from)
        except Exception as e:
            print(f"Errore formattazione date_from: {str(e)}")
    
    if date_to:
        try:
            if isinstance(date_to, (datetime, date)):
                date_to = date_to.strftime('%d/%m/%Y')
            query += " AND date <= ?"
            params.append(date_to)
        except Exception as e:
            print(f"Errore formattazione date_to: {str(e)}")
    
    query += " ORDER BY date ASC"
    
    try:
        c.execute(query, params)
        results = c.fetchall()
        
        processed_results = []
        for row in results:
            processed_results.append(dict(row))
            
        return processed_results
    except Exception as e:
        print(f"Errore recupero dati produzione: {str(e)}")
        return []
    finally:
        conn.close()

# Funzioni per calcoli e visualizzazioni
def calculate_probabilities(historical_data):
    """
    Calcola la probabilità di avere una domanda >= a un dato valore.
    Basato sui dati storici delle room nights.
    """
    # Calcola frequenze
    freq = {}
    for value in historical_data.values():
        if value in freq:
            freq[value] += 1
        else:
            freq[value] = 1
    
    # Calcola probabilità
    prob = {}
    total = len(historical_data)
    for value in sorted(freq.keys(), reverse=True):
        # Probabilità cumulativa di avere richiesta >= value
        prob[value] = sum(freq[v] for v in freq if v >= value) / total
    
    return prob

def calculate_dashboard_data(historical_data, otb_data, pickup_data, params, room_type_config=None):
    """
    Calcola i dati per la dashboard, includendo controlli di inventario e tipo di camera.
    """
    # Calcola probabilità dalla storico
    probabilities = calculate_probabilities(historical_data)
    
    # Estrai informazioni sulla tipologia di camera
    is_entry_level = params.get('is_entry_level', False)
    rooms_in_type = params.get('rooms_in_type', 0)
    if room_type_config:
        is_entry_level = room_type_config.get('is_entry_level', is_entry_level)
        rooms_in_type = room_type_config.get('rooms_in_type', rooms_in_type)
    
    dashboard_data = []
    for date, rn in otb_data.items():
        # Capacità effettiva - se disponibile
        effective_capacity = rooms_in_type
        if room_type_config and 'id' in room_type_config and 'hotel_id' in room_type_config:
            capacity = get_effective_capacity(
                room_type_config['hotel_id'],
                room_type_config['id'],
                date
            )
            if capacity > 0:
                effective_capacity = capacity
        
        # Probabilità che la domanda sia >= rn
        prob = 0
        for hist_rn, hist_prob in probabilities.items():
            if hist_rn >= rn:
                prob = hist_prob
                break
        
        # Expected Revenue
        exp_revenue = prob * params['adr']
        
        # Delta pickup
        delta = 0
        if date in pickup_data:
            delta = pickup_data[date]['2025'] - pickup_data[date]['2024']
        
        # Soglia upgrade
        threshold = params['adr'] * prob * params['min_margin']
        
        # Check se siamo in overbooking per questa tipologia
        is_overbooking = rn > effective_capacity
        
        # Decisione considerando entry-level e overbooking
        if is_entry_level or is_overbooking:
            recommendation = "No"
        else:
            recommendation = "Sì" if exp_revenue < threshold else "No"
        
        # Suggerimento personalizzato
        if is_overbooking:
            suggestion = f"Attenzione: overbooking di {rn - effective_capacity} camere"
        elif is_entry_level:
            suggestion = "Tipologia entry-level: non applicabile per upgrade"
        elif recommendation == "Sì":
            suggestion = "Bassa probabilità di vendita - considera upgrade gratuito"
        elif prob > 0.7:
            suggestion = "Alta probabilità di vendita - non fare upgrade"
        else:
            suggestion = "Probabilità media - valuta disponibilità e strategia"
        
        # Percentuale occupazione tipologia
        occupancy_pct = (rn / effective_capacity) * 100 if effective_capacity > 0 else 0
        
        dashboard_data.append({
            "Data": date,
            "RN Attuali": rn,
            "Probabilità": prob,
            "Expected Revenue": exp_revenue,
            "Delta Pickup": delta,
            "Soglia Upgrade": threshold,
            "Upgrade Consigliato": recommendation,
            "Suggerimento": suggestion,
            "% Occupazione Tipologia": occupancy_pct,
            "È Overbooking": is_overbooking,
            "Capacità Effettiva": effective_capacity
        })
    
    return dashboard_data

def create_excel_file(params, historical_data, otb_data, pickup_data, room_type_name="Standard"):
    """
    Crea un file Excel con tutte le analisi e i suggerimenti.
    """
    # Crea nuovo workbook
    wb = Workbook()
    
    # Definizione degli stili professionali
    # Colori aziendali
    color_primary = "4472C4"  # Blu aziendale
    color_secondary = "ED7D31"  # Arancione complementare
    color_accent = "70AD47"  # Verde per accenti positivi
    color_negative = "C00000"  # Rosso per valori negativi
    color_header_bg = "4472C4"  # Sfondo intestazioni
    color_header_font = "FFFFFF"  # Testo intestazioni

    # Stili di bordo
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='medium')
    )

    # Stili di font
    header_font = Font(name='Calibri', size=11, bold=True, color=color_header_font)
    title_font = Font(name='Calibri', size=14, bold=True, color=color_primary)
    normal_font = Font(name='Calibri', size=11)
    accent_font = Font(name='Calibri', size=11, color=color_accent, bold=True)
    negative_font = Font(name='Calibri', size=11, color=color_negative, bold=True)

    # Stili di riempimento
    header_fill = PatternFill(start_color=color_header_bg, end_color=color_header_bg, fill_type="solid")
    input_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Giallo chiaro più professionale
    alt_row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")  # Grigio chiaro per righe alternate

    # Allineamenti
    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    # --------------------------
    # 1. Foglio PARAMETRI
    # --------------------------
    ws_param = wb.active
    ws_param.title = "Parametri"

    # Titolo del foglio
    ws_param.merge_cells('A1:C1')
    ws_param['A1'] = "PARAMETRI DI CONFIGURAZIONE"
    ws_param['A1'].font = title_font
    ws_param['A1'].alignment = center_align

    # Intestazioni
    ws_param.append(["Parametro", "Valore", "Note"])
    for col in range(1, 4):
        cell = ws_param.cell(row=2, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = header_border
        cell.alignment = center_align

    # Dati AGGIORNATI con nuovi parametri
    parameters = [
        ["ADR Medio", params['adr'], "Tariffa media giornaliera in €"],
        ["Soglia Upgrade (fissa)", params['upgrade_threshold'], "Non utilizzata nella versione corrente"],
        ["Margine minimo richiesto", params['min_margin'], "Espresso come percentuale (0.6 = 60%)"],
        ["Giorni di analisi", params['days'], "Numero di giorni inclusi nell'analisi"],
        ["Tipologia", room_type_name, "Nome della tipologia analizzata"],
        ["È Entry-Level?", "Sì" if params['is_entry_level'] else "No", "Se Sì, questa tipologia non riceve suggerimenti di upgrade"],
        ["Camere in questa tipologia", params['rooms_in_type'], "Numero di camere disponibili per questa tipologia"],
        ["Camere totali struttura", params['total_rooms'], "Numero totale di camere nella struttura"]
    ]

    for i, (param, value, note) in enumerate(parameters, start=3):
        ws_param.cell(row=i, column=1, value=param).font = normal_font
        ws_param.cell(row=i, column=1, value=param).border = thin_border
        ws_param.cell(row=i, column=1, value=param).alignment = left_align
        
        cell = ws_param.cell(row=i, column=2, value=value)
        cell.border = thin_border
        cell.font = normal_font
        cell.alignment = right_align
        
        # Formattazione numeri in stile italiano
        if param == "ADR Medio":
            cell.number_format = '€ #.##0,00'
        elif param == "Soglia Upgrade (fissa)":
            cell.number_format = '€ #.##0,00'
        elif param == "Margine minimo richiesto":
            cell.number_format = '0%'
        
        ws_param.cell(row=i, column=3, value=note).font = normal_font
        ws_param.cell(row=i, column=3, value=note).border = thin_border
        ws_param.cell(row=i, column=3, value=note).alignment = left_align
        
        # Righe alternate per migliorare la leggibilità
        if i % 2 == 1:
            for col in range(1, 4):
                ws_param.cell(row=i, column=col).fill = alt_row_fill

    # Regolazione larghezza colonne
    ws_param.column_dimensions['A'].width = 25
    ws_param.column_dimensions['B'].width = 15
    ws_param.column_dimensions['C'].width = 40

    # --------------------------
    # 2. Foglio STORICO 2024 (unificato)
    # --------------------------
    ws_storico = wb.create_sheet("Storico 2024")

    # Titolo del foglio
    ws_storico.merge_cells('A1:E1')
    ws_storico['A1'] = "DATI STORICI GIUGNO 2024"
    ws_storico['A1'].font = title_font
    ws_storico['A1'].alignment = center_align

    # Intestazioni
    headers = ["Data", "Room Nights 2024", "Domanda", "Frequenza", "Probabilità"]
    ws_storico.append(headers)
    for col in range(1, 6):
        cell = ws_storico.cell(row=2, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = header_border
        cell.alignment = center_align

    # Dati
    for i, (date, rn) in enumerate(historical_data.items()):
        row = i + 3
        
        # Data
        date_cell = ws_storico.cell(row=row, column=1, value=date)
        date_cell.number_format = 'DD/MM/YYYY'
        date_cell.alignment = center_align
        date_cell.border = thin_border
        
        # Room Nights
        rn_cell = ws_storico.cell(row=row, column=2, value=rn)
        rn_cell.fill = input_fill
        rn_cell.border = thin_border
        rn_cell.alignment = center_align
        
        # Domanda
        dom_cell = ws_storico.cell(row=row, column=3, value=f"=B{row}")
        dom_cell.border = thin_border
        dom_cell.alignment = center_align
        
        # Frequenza
        freq_cell = ws_storico.cell(row=row, column=4, value=f"=CONTA.SE(C$3:C$32;C{row})")
        freq_cell.border = thin_border
        freq_cell.alignment = center_align
        
        # Probabilità
        prob_cell = ws_storico.cell(row=row, column=5, value=f"=D{row}/CONTA.VALORI(C$3:C$32)")
        prob_cell.number_format = '0,00%'
        prob_cell.border = thin_border
        prob_cell.alignment = center_align
        
        # Righe alternate
        if i % 2 == 1:
            for col in range(1, 6):
                ws_storico.cell(row=row, column=col).fill = alt_row_fill

    # --------------------------
    # 3. Foglio OTB GIUGNO 2025
    # --------------------------
    ws_otb = wb.create_sheet("OTB Giugno 2025")

    # Titolo del foglio
    ws_otb.merge_cells('A1:C1')
    ws_otb['A1'] = "PRENOTAZIONI ACQUISITE (ON THE BOOKS) - GIUGNO 2025"
    ws_otb['A1'].font = title_font
    ws_otb['A1'].alignment = center_align

    # Intestazioni
    headers = ["Data", "Room Nights", "Note"]
    ws_otb.append(headers)
    for col in range(1, 4):
        cell = ws_otb.cell(row=2, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = header_border
        cell.alignment = center_align

    # Dati
    for i, (date, rn) in enumerate(otb_data.items()):
        row = i + 3
        
        # Data
        date_cell = ws_otb.cell(row=row, column=1, value=date)
        date_cell.number_format = 'DD/MM/YYYY'
        date_cell.alignment = center_align
        date_cell.border = thin_border
        
        # Room Nights
        rn_cell = ws_otb.cell(row=row, column=2, value=rn)
        rn_cell.fill = input_fill
        rn_cell.border = thin_border
        rn_cell.alignment = center_align
        
        # Note (cella vuota formattata)
        note_cell = ws_otb.cell(row=row, column=3)
        note_cell.border = thin_border
        
        # Righe alternate
        if i % 2 == 1:
            for col in range(1, 4):
                ws_otb.cell(row=row, column=col).fill = alt_row_fill

    # --------------------------
    # 4. Foglio PICKUP vs SPIT
    # --------------------------
    ws_pickup = wb.create_sheet("Pickup vs SPIT")

    # Titolo del foglio
    ws_pickup.merge_cells('A1:D1')
    ws_pickup['A1'] = "CONFRONTO PICKUP A 7 GIORNI: 2025 vs 2024"
    ws_pickup['A1'].font = title_font
    ws_pickup['A1'].alignment = center_align

    # Intestazioni
    headers = ["Data", "Pickup 7gg 2025", "Pickup 7gg 2024", "Delta"]
    ws_pickup.append(headers)
    for col in range(1, 5):
        cell = ws_pickup.cell(row=2, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = header_border
        cell.alignment = center_align

    # Dati
    for i, (date, data) in enumerate(pickup_data.items()):
        row = i + 3
        
        # Data
        date_cell = ws_pickup.cell(row=row, column=1, value=date)
        date_cell.number_format = 'DD/MM/YYYY'
        date_cell.alignment = center_align
        date_cell.border = thin_border
        
        # Pickup 2025
        pickup25_cell = ws_pickup.cell(row=row, column=2, value=data['2025'])
        pickup25_cell.fill = input_fill
        pickup25_cell.border = thin_border
        pickup25_cell.alignment = center_align
        
        # Pickup 2024
        pickup24_cell = ws_pickup.cell(row=row, column=3, value=data['2024'])
        pickup24_cell.fill = input_fill
        pickup24_cell.border = thin_border
        pickup24_cell.alignment = center_align
        
        # Delta
        delta_cell = ws_pickup.cell(row=row, column=4, value=f"=B{row}-C{row}")
        delta_cell.border = thin_border
        delta_cell.alignment = center_align
        
        # Righe alternate
        if i % 2 == 1:
            for col in range(1, 5):
                ws_pickup.cell(row=row, column=col).fill = alt_row_fill

    # Formattazione condizionale su Delta
    ws_pickup.conditional_formatting.add(
        f"D3:D{row}", CellIsRule(operator="lessThan", formula=["0"], 
                                stopIfTrue=False, 
                                fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"), 
                                font=Font(color="9C0006"))
    )
    ws_pickup.conditional_formatting.add(
        f"D3:D{row}", CellIsRule(operator="greaterThan", formula=["0"], 
                                stopIfTrue=False, 
                                fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"), 
                                font=Font(color="006100"))
    )

    # --------------------------
    # 5. Foglio DASHBOARD
    # --------------------------
    ws_dash = wb.create_sheet("Dashboard Upgrade")

    # Titolo del foglio
    ws_dash.merge_cells('A1:I1')
    ws_dash['A1'] = f"DASHBOARD DECISIONALE UPGRADE - {room_type_name} - GIUGNO 2025"
    ws_dash['A1'].font = title_font
    ws_dash['A1'].alignment = center_align

    # Intestazioni - AGGIUNTE COLONNE
    headers = [
        "Data", "RN Attuali", "Prob. Domanda ≥ RN", "Expected Revenue",
        "Delta Pickup vs LY", "Soglia Upgrade Dinamica (€)",
        "% Occupazione", "Upgrade Consigliato", "Suggerimento Operativo"
    ]
    ws_dash.append(headers)
    for col in range(1, 10):
        cell = ws_dash.cell(row=2, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = header_border
        cell.alignment = center_align

    # Dati dashboard
    dates = list(otb_data.keys())
    for i, date in enumerate(dates):
        row = i + 3
        
        # Data
        date_cell = ws_dash.cell(row=row, column=1, value=date)
        date_cell.number_format = 'DD/MM/YYYY'
        date_cell.alignment = center_align
        date_cell.border = thin_border
        
        # RN Attuali
        rn_cell = ws_dash.cell(row=row, column=2, value=f'=CERCA.X(A{row};\'OTB Giugno 2025\'!A:A;\'OTB Giugno 2025\'!B:B;"")')
        rn_cell.border = thin_border
        rn_cell.alignment = center_align
        
        # Probabilità
        prob_cell = ws_dash.cell(row=row, column=3, value=f'=SOMMA.SE(\'Storico 2024\'!C$3:C$32;">=" & B{row};\'Storico 2024\'!E$3:E$32)')
        prob_cell.number_format = '0,00%'
        prob_cell.border = thin_border
        prob_cell.alignment = center_align
        
        # Expected Revenue
        rev_cell = ws_dash.cell(row=row, column=4, value=f'=C{row}*Parametri!$B$2')
        rev_cell.number_format = '€ #.##0,00'
        rev_cell.border = thin_border
        rev_cell.alignment = center_align
        
        # Delta Pickup
        delta_cell = ws_dash.cell(row=row, column=5, value=f'=CERCA.X(A{row};\'Pickup vs SPIT\'!A:A;\'Pickup vs SPIT\'!D:D;"")')
        delta_cell.border = thin_border
        delta_cell.alignment = center_align
        
        # Soglia Upgrade
        soglia_cell = ws_dash.cell(row=row, column=6, value=f'=Parametri!$B$2*C{row}*Parametri!$B$3')
        soglia_cell.number_format = '€ #.##0,00'
        soglia_cell.border = thin_border
        soglia_cell.alignment = center_align
        
        # % Occupazione - NUOVO
        occ_cell = ws_dash.cell(row=row, column=7, value=f'=B{row}/Parametri!$B$7*100')
        occ_cell.number_format = '0,00%'
        occ_cell.border = thin_border
        occ_cell.alignment = center_align
        
        # Upgrade Consigliato - AGGIORNATO CON CONTROLLI E FORMULA ITALIANA
        upgrade_cell = ws_dash.cell(row=row, column=8, value=
            f'=SE(O(Parametri!$B$6="Sì"; B{row}>Parametri!$B$7); "No"; SE(D{row}<F{row}; "Sì"; "No"))')
        upgrade_cell.border = thin_border
        upgrade_cell.alignment = center_align
        
        # Suggerimento - AGGIORNATO CON CONTROLLI E FORMULA ITALIANA
        sugg_cell = ws_dash.cell(
            row=row, column=9, 
            value=(
                f'=SE(B{row}>Parametri!$B$7;"Attenzione: overbooking di " & TESTO(B{row}-Parametri!$B$7;"0") & " camere";'
                f'SE(Parametri!$B$6="Sì";"Tipologia entry-level: non applicabile per upgrade";'
                f'SE(H{row}="Sì";'
                f'"Bassa probabilità di vendita - considera upgrade gratuito";'
                f'SE(C{row}>0,7;"Alta probabilità di vendita - non fare upgrade";'
                f'"Probabilità media - valuta disponibilità e strategia"))))'
            )
        )
        sugg_cell.border = thin_border
        sugg_cell.alignment = left_align
        
        # Righe alternate
        if i % 2 == 1:
            for col in range(1, 10):
                ws_dash.cell(row=row, column=col).fill = alt_row_fill

    # Formattazioni condizionali per Upgrade
    ws_dash.conditional_formatting.add(
        f"H3:H{row}", CellIsRule(operator="equal", formula=['"Sì"'], 
                                stopIfTrue=False, 
                                fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"), 
                                font=Font(color="006100"))
    )
    ws_dash.conditional_formatting.add(
        f"H3:H{row}", CellIsRule(operator="equal", formula=['"No"'], 
                                stopIfTrue=False, 
                                fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"), 
                                font=Font(color="9C0006"))
    )
    
    # Formattazione per overbooking - NUOVO
    ws_dash.conditional_formatting.add(
        f"G3:G{row}", CellIsRule(operator="greaterThan", formula=["1"], 
                               stopIfTrue=False, 
                               fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"), 
                               font=Font(color="9C0006"))
    )

    # Formattazione scala colore per probabilità
    ws_dash.conditional_formatting.add(
        f"C3:C{row}", ColorScaleRule(start_type='min', start_color='F8696B',
                                   mid_type='percentile', mid_value=50, mid_color='FFEB84',
                                   end_type='max', end_color='63BE7B')
    )

    # --------------------------
    # 6. Foglio STATO INVENTARIO
    # --------------------------
    ws_inventory = wb.create_sheet("Stato Inventario")

    # Titolo del foglio
    ws_inventory.merge_cells('A1:F1')
    ws_inventory['A1'] = f"STATO INVENTARIO - {room_type_name} - GIUGNO 2025"
    ws_inventory['A1'].font = title_font
    ws_inventory['A1'].alignment = center_align

    # Intestazioni
    headers = [
        "Data", "RN Attuali", "Camere Disponibili", "Camere Rimanenti", 
        "% Occupazione", "Stato"
    ]
    ws_inventory.append(headers)
    for col in range(1, 7):
        cell = ws_inventory.cell(row=2, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = header_border
        cell.alignment = center_align

    # Popola dati inventario
    for i, (date, rn) in enumerate(otb_data.items()):
        row = i + 3
        
        # Data
        date_cell = ws_inventory.cell(row=row, column=1, value=date)
        date_cell.number_format = 'DD/MM/YYYY'
        date_cell.alignment = center_align
        date_cell.border = thin_border
        
        # RN Attuali
        rn_cell = ws_inventory.cell(row=row, column=2, value=rn)
        rn_cell.border = thin_border
        rn_cell.alignment = center_align
        
        # Camere Disponibili
        avail_cell = ws_inventory.cell(row=row, column=3, value=params['rooms_in_type'])
        avail_cell.border = thin_border
        avail_cell.alignment = center_align
        
        # Camere Rimanenti
        remain_cell = ws_inventory.cell(row=row, column=4, value=f"=C{row}-B{row}")
        remain_cell.border = thin_border
        remain_cell.alignment = center_align
        
        # % Occupazione
        occ_cell = ws_inventory.cell(row=row, column=5, value=f"=B{row}/C{row}")
        occ_cell.number_format = '0.00%'
        occ_cell.border = thin_border
        occ_cell.alignment = center_align
        
        # Stato
        status_cell = ws_inventory.cell(
            row=row, column=6, 
            value=f'=SE(D{row}<0;"OVERBOOKING";SE(D{row}=0;"TUTTO VENDUTO";SE(D{row}/C{row}<0.2;"QUASI ESAURITO";"DISPONIBILE")))'
        )
        status_cell.border = thin_border
        status_cell.alignment = center_align
        
        # Righe alternate
        if i % 2 == 1:
            for col in range(1, 7):
                ws_inventory.cell(row=row, column=col).fill = alt_row_fill

    # Formattazione condizionale per stato
    ws_inventory.conditional_formatting.add(
        f"F3:F{row}", CellIsRule(operator="equal", formula=['"OVERBOOKING"'], 
                               stopIfTrue=False, 
                               fill=PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"), 
                               font=Font(color="FFFFFF"))
    )
    
    ws_inventory.conditional_formatting.add(
        f"F3:F{row}", CellIsRule(operator="equal", formula=['"TUTTO VENDUTO"'], 
                               stopIfTrue=False, 
                               fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"), 
                               font=Font(color="9C0006"))
    )
    
    ws_inventory.conditional_formatting.add(
        f"F3:F{row}", CellIsRule(operator="equal", formula=['"QUASI ESAURITO"'], 
                               stopIfTrue=False, 
                               fill=PatternFill(start_color="FFEB84", end_color="FFEB84", fill_type="solid"), 
                               font=Font(color="9C5700"))
    )
    
    ws_inventory.conditional_formatting.add(
        f"F3:F{row}", CellIsRule(operator="equal", formula=['"DISPONIBILE"'], 
                               stopIfTrue=False, 
                               fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"), 
                               font=Font(color="006100"))
    )
    
    # Imposta larghezza colonne
    for col in range(1, 7):
        ws_inventory.column_dimensions[get_column_letter(col)].width = 18

    # Imposta dashboard come foglio attivo
    wb.active = wb["Dashboard Upgrade"]
    
    # Salva in buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer

# Funzione di configurazione delle tipologie del Cala Cuncheddi
def setup_cala_cuncheddi_presets():
    """
    Configura automaticamente le tipologie di camera del Cala Cuncheddi
    con le relative gerarchie di upgrade.
    """
    # Verifica se l'hotel esiste già
    hotel = get_hotel_by_name("Cala Cuncheddi")
    if not hotel:
        # Crea l'hotel se non esiste
        hotel_id = add_hotel({
            'hotel_name': "Cala Cuncheddi",
            'total_rooms': 85,
            'address': "Via Dei Ginepri 3, 07026 Pittulongu, Olbia",
            'city': "Olbia",
            'stars': 5,
            'seasonal': True  # Hotel stagionale
        })
    else:
        hotel_id = hotel['id']
    
    # Definisci le tipologie di camera con le quantità aggiornate
    room_types = [
        {
            'name': "Classic Garden",
            'count': 33,
            'is_entry_level': True,
            'adr': 300,
            'upgrade_to': "Classic Sea View"
        },
        {
            'name': "Classic Sea View",
            'count': 12,
            'is_entry_level': False,
            'adr': 350,
            'upgrade_to': "Superior Sea View"
        },
        {
            'name': "Family",
            'count': 2,
            'is_entry_level': False,
            'adr': 495,
            'upgrade_to': "Junior Suite Pool"
        },
        {
            'name': "Superior Sea View",
            'count': 18,
            'is_entry_level': False,
            'adr': 400,
            'upgrade_to': "Executive"
        },
        {
            'name': "Executive",
            'count': 9,
            'is_entry_level': False,
            'adr': 450,
            'upgrade_to': "Deluxe"
        },
        {
            'name': "Deluxe",
            'count': 4,
            'is_entry_level': False,
            'adr': 520,
            'upgrade_to': "Junior Suite Pool"
        },
        {
            'name': "Junior Suite Pool",
            'count': 4,
            'is_entry_level': False,
            'adr': 650,
            'upgrade_to': "Suite"
        },
        {
            'name': "Suite",
            'count': 3,
            'is_entry_level': False,
            'adr': 780,
            'upgrade_to': None
        }
    ]
    
    # Aggiungi/aggiorna tutte le tipologie
    for rt in room_types:
        room_type = get_room_type_by_name(hotel_id, rt['name'])
        if room_type:
            # Aggiorna la tipologia esistente
            update_room_type(room_type['id'], {
                'room_type_name': rt['name'],
                'rooms_in_type': rt['count'],
                'is_entry_level': rt['is_entry_level'],
                'adr': rt['adr'],
                'upgrade_target_type': rt['upgrade_to']
            })
        else:
            # Crea nuova tipologia
            add_room_type(hotel_id, {
                'room_type_name': rt['name'],
                'rooms_in_type': rt['count'],
                'is_entry_level': rt['is_entry_level'],
                'adr': rt['adr'],
                'upgrade_target_type': rt['upgrade_to'],
                'min_margin': 0.6  # Margine predefinito
            })
    
    return True

# Layout dell'interfaccia principale
st.title("🏨 Hotel Upgrade Advisor Pro")
st.write("Sistema avanzato per decisioni di revenue management con controllo inventario")

# Inizializza il database
init_database()

# Sidebar per configurazione
st.sidebar.header("Configurazione")

# Verifica se esiste una configurazione per Cala Cuncheddi
hotel = get_hotel_by_name("Cala Cuncheddi")
if not hotel:
    if st.sidebar.button("Configura Cala Cuncheddi"):
        if setup_cala_cuncheddi_presets():
            st.sidebar.success("Cala Cuncheddi configurato con successo!")
            hotel = get_hotel_by_name("Cala Cuncheddi")
        else:
            st.sidebar.error("Errore nella configurazione di Cala Cuncheddi")

# Selezione hotel
hotels = get_all_hotels()
if hotels:
    hotel_names = [h['hotel_name'] for h in hotels]
    default_index = hotel_names.index("Cala Cuncheddi") if "Cala Cuncheddi" in hotel_names else 0
    selected_hotel_name = st.sidebar.selectbox("Seleziona Hotel", hotel_names, index=default_index)
    selected_hotel = get_hotel_by_name(selected_hotel_name)
else:
    st.sidebar.warning("Nessun hotel configurato")
    selected_hotel_name = "Hotel Demo"
    selected_hotel = None

# Selezione tipologia camera
room_types = []
if selected_hotel:
    room_types = get_room_types_by_hotel(selected_hotel['id'])
    if room_types:
        room_type_names = [rt['room_type_name'] for rt in room_types]
        selected_room_type_name = st.sidebar.selectbox("Seleziona Tipologia", room_type_names)
        selected_room_type = next((rt for rt in room_types if rt['room_type_name'] == selected_room_type_name), None)
    else:
        st.sidebar.warning(f"Nessuna tipologia configurata per {selected_hotel_name}")
        selected_room_type_name = "Standard"
        selected_room_type = None
else:
    selected_room_type_name = "Standard"
    selected_room_type = None

# Parametri di base
with st.sidebar.expander("Parametri di Base", expanded=True):
    # Se abbiamo un hotel e una tipologia selezionati, usa i valori configurati
    if selected_room_type:
        adr = st.number_input("ADR Medio (€)", min_value=50.0, max_value=1000.0, value=selected_room_type.get('adr', 300.0), step=5.0)
        min_margin = st.slider("Margine minimo richiesto", min_value=0.1, max_value=0.9, 
                              value=selected_room_type.get('min_margin', 0.6), step=0.05, format="%.2f")
        upgrade_threshold = st.number_input("Soglia Upgrade fissa (€)", min_value=0.0, max_value=500.0, 
                                          value=selected_room_type.get('upgrade_threshold', 100.0), step=10.0)
        days = st.slider("Giorni di analisi", min_value=7, max_value=60, value=30, step=1)
        
        # Parametri inventario direttamente dalla configurazione salvata
        is_entry_level = selected_room_type.get('is_entry_level', False)
        rooms_in_type = selected_room_type.get('rooms_in_type', 20)
        st.write(f"**Camere in questa tipologia:** {rooms_in_type}")
        st.write(f"**Tipologia entry-level:** {'Sì' if is_entry_level else 'No'}")
        st.write(f"**Totale camere hotel:** {selected_hotel.get('total_rooms', 0)}")
    else:
        # Valori di default
        adr = st.number_input("ADR Medio (€)", min_value=50.0, max_value=1000.0, value=300.0, step=5.0)
        min_margin = st.slider("Margine minimo richiesto", min_value=0.1, max_value=0.9, value=0.6, step=0.05, format="%.2f")
        upgrade_threshold = st.number_input("Soglia Upgrade fissa (€)", min_value=0.0, max_value=500.0, value=100.0, step=10.0)
        days = st.slider("Giorni di analisi", min_value=7, max_value=60, value=30, step=1)
        
        # Parametri inventario manuali
        st.markdown("### Configurazione Inventario")
        is_entry_level = st.checkbox("È tipologia entry-level?", value=False)
        rooms_in_type = st.number_input("Camere in questa tipologia", min_value=1, max_value=1000, value=20, step=1)
        total_rooms = st.number_input("Camere totali struttura", min_value=1, max_value=5000, value=85, step=1)

# Pulsante per salvare configurazione
if selected_hotel and selected_room_type:
    if st.sidebar.button("Aggiorna Configurazione"):
        updated = update_room_type(selected_room_type['id'], {
            'room_type_name': selected_room_type_name,
            'rooms_in_type': rooms_in_type,
            'is_entry_level': is_entry_level,
            'adr': adr,
            'min_margin': min_margin,
            'upgrade_threshold': upgrade_threshold
        })
        if updated:
            st.sidebar.success(f"Configurazione per {selected_room_type_name} aggiornata!")
        else:
            st.sidebar.error("Errore nell'aggiornamento della configurazione")

# Prepara parametri
params = {
    'adr': adr,
    'min_margin': min_margin,
    'upgrade_threshold': upgrade_threshold,
    'days': days,
    'is_entry_level': is_entry_level,
    'room_type_name': selected_room_type_name,
    'rooms_in_type': rooms_in_type,
    'total_rooms': selected_hotel.get('total_rooms', total_rooms) if selected_hotel else total_rooms
}

# Opzioni di caricamento dati
st.sidebar.header("Caricamento Dati")
data_option = st.sidebar.radio("Seleziona metodo di caricamento", 
                              ["Carica dati da BI", "Carica file Excel", "Genera dati casuali", "Inserisci manualmente", "Usa dati salvati"])

# Inizializza dati
historical_data = {}
otb_data = {}
pickup_data = {}

# Opzione 1: Carica dati da BI
if data_option == "Carica dati da BI":
    st.sidebar.subheader("Importazione dati da Business Intelligence")
    
    # Caricamento file produzione
    st.sidebar.markdown("### File di Produzione Giornaliera")
    production_file = st.sidebar.file_uploader("Carica file produzione", type=['xlsx', 'xls'])
    
    # Caricamento file pickup
    st.sidebar.markdown("### File di Pickup")
    pickup_file = st.sidebar.file_uploader("Carica file pickup", type=['xlsx', 'xls'])
    
    # Estrazione automatica metadati
    auto_detected = False
    if production_file:
        with st.spinner("Analisi del file in corso..."):
            metadata = extract_room_type_from_powerbi(production_file, selected_hotel_name if selected_hotel else "Cala Cuncheddi")
            
            if metadata and metadata['hotel_name'] and metadata['room_type']:
                auto_detected = True
                st.success(f"Rilevato automaticamente: Hotel {metadata['hotel_name']}, Tipologia {metadata['room_type']}")
                
                # Ottieni configurazioni hotel se esistenti
                hotel_config = get_hotel_by_name(metadata['hotel_name'])
                if not hotel_config:
                    st.info(f"Hotel {metadata['hotel_name']} non trovato nella configurazione. Verrà creato automaticamente.")
                    
                    # Crea hotel se non esiste
                    hotel_id = add_hotel({
                        'hotel_name': metadata['hotel_name'],
                        'total_rooms': 85,  # Default per Cala Cuncheddi
                        'city': "Olbia" if "cala" in metadata['hotel_name'].lower() else ""
                    })
                else:
                    hotel_id = hotel_config['id']
                
                # Ottieni configurazione tipologia se esiste
                room_type_config = get_room_type_by_name(hotel_id, metadata['room_type'])
                if not room_type_config:
                    st.info(f"Tipologia {metadata['room_type']} non trovata nella configurazione. Verrà creata automaticamente.")
                    
                    # Determina se potrebbe essere entry-level
                    is_entry = "classic garden" in metadata['room_type'].lower()
                    
                    # Determina capacità basandosi sulle tipologie del Cala Cuncheddi
                    capacity_map = {
                        "Classic Garden": 33,
                        "Classic Sea View": 12,
                        "Family": 2,
                        "Superior Sea View": 18,
                        "Executive": 9,
                        "Deluxe": 4,
                        "Junior Suite Pool": 4,
                        "Suite": 3
                    }
                    capacity = capacity_map.get(metadata['room_type'], 10)
                    
                    # Crea tipologia se non esiste
                    room_type_id = add_room_type(hotel_id, {
                        'room_type_name': metadata['room_type'],
                        'rooms_in_type': capacity,
                        'is_entry_level': is_entry,
                        'adr': 300.0  # Default
                    })
    
    # Continua con il normale processo di elaborazione
    if st.sidebar.button("Elabora dati BI"):
        if production_file is None and pickup_file is None:
            st.sidebar.error("Carica almeno uno dei file richiesti.")
        else:
            daily_production = None
            pickup_entries = None
            
            # Elabora file produzione
            if production_file:
                with st.spinner("Elaborazione file produzione..."):
                    daily_production, prod_error = process_daily_production_excel(production_file, selected_hotel_name)
                    if prod_error:
                        st.error(f"Errore nel file di produzione: {prod_error}")
                    elif daily_production:
                        # Salva i dati di produzione
                        saved_count = save_production_data(selected_hotel_name, daily_production)
                        st.success(f"File produzione elaborato: {len(daily_production)} giorni trovati, {saved_count} salvati")
                        st.dataframe(pd.DataFrame(daily_production).head())
            
            # Elabora file pickup
            if pickup_file:
                with st.spinner("Elaborazione file pickup..."):
                    pickup_entries, pickup_error = process_pickup_excel(pickup_file, selected_hotel_name)
                    if pickup_error:
                        st.error(f"Errore nel file pickup: {pickup_error}")
                    elif pickup_entries:
                        # Salva i dati di pickup
                        saved_count = save_pickup_data(selected_hotel_name, pickup_entries)
                        st.success(f"File pickup elaborato: {len(pickup_entries)} giorni trovati, {saved_count} salvati")
                        st.dataframe(pd.DataFrame(pickup_entries).head())
            
            # Mappa i dati al modello
            if daily_production or pickup_entries:
                historical_data, otb_data, pickup_data = map_bi_data_to_model(daily_production, pickup_entries)
                
                if historical_data:
                    st.success(f"Dati storici mappati: {len(historical_data)} giorni")
                if otb_data:
                    st.success(f"Dati OTB mappati: {len(otb_data)} giorni")
                if pickup_data:
                    st.success(f"Dati pickup mappati: {len(pickup_data)} giorni")
    
    # Analisi trend vs SPIT
    if st.sidebar.checkbox("Mostra analisi trend vs SPIT"):
        trend_days = st.sidebar.slider("Giorni da analizzare", 7, 180, 60)
        trend_data = analyze_vs_spit_trend(selected_hotel_name, days=trend_days)
        
        if "error" in trend_data:
            st.sidebar.error(trend_data["error"])
        else:
            st.sidebar.success(f"Analisi ultimi {trend_data['period_days']} giorni:")
            st.sidebar.metric("Room Nights Totali", f"{trend_data['total_rn']:.0f}")
            st.sidebar.metric("vs SPIT", f"{trend_data['total_vs_spit']:.0f} ({trend_data['vs_spit_pct']:.1f}%)")
            st.sidebar.metric("Room Nights Anno Precedente", f"{trend_data['spit_rn']:.0f}")

# Opzione 2: Carica file Excel (invariata)
# Opzione 3: Genera dati casuali (invariata)
# Opzione 4: Inserimento manuale (invariata)
# Opzione 5: Usa dati salvati (invariata)

# Assicurarsi che ci siano dati da mostrare
if not historical_data or not otb_data:
    st.warning("Per favore, carica o genera i dati da analizzare")
else:
    # Calcola i dati della dashboard con la configurazione della tipologia
    dashboard_data = calculate_dashboard_data(historical_data, otb_data, pickup_data, params, selected_room_type)
    
    # Visualizza i risultati
    tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs([
        "Dashboard", "Stato Inventario", "Grafici", "Dati di Input", 
        "Previsioni", "Visione Complessiva"
    ])
    
    # Tab 1: Dashboard
    with tab1:
        st.subheader(f"Dashboard Decisionale Upgrade - {selected_room_type_name}")
        
        # Visualizza avviso se è tipologia entry-level
        if is_entry_level:
            st.warning("⚠️ Questa è una tipologia entry-level. Non sono consigliati upgrade gratuiti da questa categoria.")
        
        # Metriche principali
        col1, col2, col3, col4 = st.columns(4)
        
        # Calcola metriche chiave
        total_rooms_booked = sum(otb.get("RN Attuali", 0) for otb in dashboard_data)
        avg_probability = sum(d.get("Probabilità", 0) for d in dashboard_data) / len(dashboard_data) if dashboard_data else 0
        recommended_upgrades = sum(1 for d in dashboard_data if d.get("Upgrade Consigliato") == "Sì")
        
        # Calcola overbooking e percentuale di occupazione
        overbooking_days = sum(1 for d in dashboard_data if d.get("È Overbooking", False))
        avg_occupancy = sum(d.get("% Occupazione Tipologia", 0) for d in dashboard_data) / len(dashboard_data) if dashboard_data else 0
        
        with col1:
            st.metric(label="Room Nights Totali", value=f"{total_rooms_booked}")
        with col2:
            st.metric(label=f"Occupazione Media {selected_room_type_name}", value=f"{avg_occupancy:.1f}%")
        with col3:
            st.metric(label="Upgrade Consigliati", value=f"{recommended_upgrades}")
        with col4:
            # Mostra giorni in overbooking se presenti
            if overbooking_days > 0:
                st.metric(label="Giorni in Overbooking", value=f"{overbooking_days}", delta=f"{overbooking_days}", delta_color="inverse")
            else:
                st.metric(label="Probabilità Media", value=f"{avg_probability:.1%}")
        
        # Tabella con i risultati della dashboard
        df_dashboard = pd.DataFrame(dashboard_data)
        
        # Formatta colonne
        df_dashboard["Data"] = df_dashboard["Data"].dt.strftime("%d/%m/%Y")
        df_dashboard["Probabilità"] = df_dashboard["Probabilità"].map("{:.1%}".format)
        df_dashboard["Expected Revenue"] = df_dashboard["Expected Revenue"].map("€{:,.2f}".format).str.replace(".", "X").str.replace(",", ".").str.replace("X", ",")
        df_dashboard["Soglia Upgrade"] = df_dashboard["Soglia Upgrade"].map("€{:,.2f}".format).str.replace(".", "X").str.replace(",", ".").str.replace("X", ",")
        df_dashboard["% Occupazione Tipologia"] = df_dashboard["% Occupazione Tipologia"].map("{:.1f}%".format)
        
        # Colonne da visualizzare
        display_columns = ["Data", "RN Attuali", "% Occupazione Tipologia", "Probabilità", "Expected Revenue", "Soglia Upgrade", "Upgrade Consigliato", "Suggerimento"]
        
        # Crea dataframe formattato
        st.dataframe(df_dashboard[display_columns], use_container_width=True)
        
        # Avvisi di stato inventario
        if overbooking_days > 0:
            st.error(f"⚠️ ATTENZIONE: Rilevato overbooking in {overbooking_days} giorni. Verificare la scheda 'Stato Inventario'.")
        
        # Download del file Excel
        if st.button("Genera File Excel"):
            excel_buffer = create_excel_file(params, historical_data, otb_data, pickup_data, selected_room_type_name)
            
            st.download_button(
                label="📥 Scarica Excel",
                data=excel_buffer,
                file_name=f"hotel_upgrade_advisor_{selected_hotel_name}_{selected_room_type_name}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    
    # Tab 2: Stato Inventario
    with tab2:
        st.subheader(f"Stato Inventario - {selected_room_type_name}")
        
        # Informazioni generali
        st.info(f"""
        **Informazioni Tipologia**
        - Nome: {selected_room_type_name}
        - Camere disponibili: {rooms_in_type} su {params['total_rooms']} totali ({(rooms_in_type/params['total_rooms']*100):.1f}% della struttura)
        - Tipologia entry-level: {"Sì" if is_entry_level else "No"}
        """)
        
        # Crea visualizzazione stato inventario
        inventory_data = []
        for item in dashboard_data:
            date = item["Data"]
            rn = item["RN Attuali"]
            occupancy_pct = item["% Occupazione Tipologia"]
            rooms_remaining = rooms_in_type - rn
            
            if rooms_remaining < 0:
                status = "OVERBOOKING"
                status_color = "red"
            elif rooms_remaining == 0:
                status = "TUTTO VENDUTO"
                status_color = "orange"
            elif rooms_remaining / rooms_in_type < 0.2:
                status = "QUASI ESAURITO"
                status_color = "yellow"
            else:
                status = "DISPONIBILE"
                status_color = "green"
            
            inventory_data.append({
                "Data": date,
                "Room Nights": rn,
                "Camere Disponibili": rooms_in_type,
                "Camere Rimanenti": rooms_remaining,
                "% Occupazione": occupancy_pct,
                "Stato": status,
                "Colore": status_color
            })
        
        df_inventory = pd.DataFrame(inventory_data)
        
        # Formatta date e percentuali
        df_inventory["Data"] = df_inventory["Data"].dt.strftime("%d/%m/%Y")
        df_inventory["% Occupazione"] = df_inventory["% Occupazione"].map("{:.1f}%".format)
        
        # Applica colori alle celle della colonna Stato
        def color_status(val):
            if val == "OVERBOOKING":
                return 'background-color: red; color: white'
            elif val == "TUTTO VENDUTO":
                return 'background-color: orange; color: white'
            elif val == "QUASI ESAURITO":
                return 'background-color: #FFEB84; color: black'
            else:
                return 'background-color: #C6EFCE; color: black'
        
        # Visualizzazione tabella colorata
        st.dataframe(df_inventory.style.applymap(color_status, subset=['Stato']), use_container_width=True)
        
        # Grafico andamento inventario
        st.subheader("Andamento Occupazione")
        df_inventory_chart = pd.DataFrame(inventory_data)
        
        fig_inventory = px.line(
            df_inventory_chart, 
            x="Data", 
            y="Room Nights",
            title=f"Andamento Occupazione {selected_room_type_name}",
            labels={"Room Nights": "Camere Vendute", "Data": "Data"}
        )
        
        # Aggiungi linea di capacità massima
        fig_inventory.add_hline(
            y=rooms_in_type, 
            line_dash="dash", 
            line_color="red",
            annotation_text=f"Capacità Max ({rooms_in_type})",
            annotation_position="top right"
        )
        
        st.plotly_chart(fig_inventory, use_container_width=True)
        
        # Riepilogo occupazione
        st.subheader("Riepilogo Stato Occupazione")
        
        status_counts = df_inventory["Stato"].value_counts().reset_index()
        status_counts.columns = ["Stato", "Giorni"]
        
        fig_status = px.pie(
            status_counts, 
            names="Stato", 
            values="Giorni",
            color="Stato",
            color_discrete_map={
                "DISPONIBILE": "#C6EFCE",
                "QUASI ESAURITO": "#FFEB84",
                "TUTTO VENDUTO": "orange",
                "OVERBOOKING": "red"
            },
            title="Distribuzione Stato Occupazione"
        )
        
        st.plotly_chart(fig_status, use_container_width=True)

    # Tab 3: Grafici (invariata)
    # Tab 4: Dati di Input (invariata) 
    # Tab 5: Previsioni (invariata)
    
    # Tab 6: Visione Complessiva (nuova)
    with tab6:
        if selected_hotel:
            st.subheader(f"Visione Complessiva {selected_hotel_name}")
            
            # Recupera tutte le tipologie dell'hotel
            all_room_types = get_room_types_by_hotel(selected_hotel['id'])
            
            if all_room_types:
                # Visualizzazione grafica distribuzione camere
                room_counts = [rt['rooms_in_type'] for rt in all_room_types]
                room_names = [rt['room_type_name'] for rt in all_room_types]
                
                fig_rooms = px.pie(
                    values=room_counts,
                    names=room_names,
                    title=f"Distribuzione Camere {selected_hotel_name}"
                )
                st.plotly_chart(fig_rooms, use_container_width=True)
                
                # Tabella riepilogativa tipologie
                st.subheader("Tutte le Tipologie")
                
                types_data = []
                for rt in all_room_types:
                    # Evidenzia la tipologia selezionata
                    is_selected = rt['room_type_name'] == selected_room_type_name
                    
                    types_data.append({
                        "Tipologia": rt['room_type_name'],
                        "Camere": rt['rooms_in_type'],
                        "% su Totale": f"{(rt['rooms_in_type'] / selected_hotel['total_rooms'] * 100):.1f}%",
                        "ADR": f"€{rt['adr']:.2f}".replace(".", ","),
                        "Entry-Level": "✓" if rt['is_entry_level'] else "",
                        "Upgrade a": rt.get('upgrade_target_type', ""),
                        "Tipologia Attuale": "✓" if is_selected else ""
                    })
                
                # Crea dataframe
                df_types = pd.DataFrame(types_data)
                
                # Applica stile per evidenziare la tipologia selezionata
                def highlight_selected(s):
                    return ['background-color: #e6f7ff' if s.name == selected_room_type_name else '' for _ in s]
                
                st.dataframe(df_types, use_container_width=True)
                
                # Visualizzazione gerarchia di upgrade
                st.subheader("Gerarchia di Upgrade")
                
                try:
                    # Costruisci grafo delle relazioni di upgrade
                    import networkx as nx
                    import matplotlib.pyplot as plt
                    
                    G = nx.DiGraph()
                    
                    # Aggiungi nodi e archi
                    for rt in all_room_types:
                        is_selected = rt['room_type_name'] == selected_room_type_name
                        G.add_node(rt['room_type_name'], 
                                  count=rt['rooms_in_type'], 
                                  adr=rt['adr'],
                                  is_entry=rt['is_entry_level'],
                                  is_selected=is_selected)
                        
                        if rt.get('upgrade_target_type'):
                            G.add_edge(rt['room_type_name'], rt['upgrade_target_type'])
                    
                    # Visualizza grafo
                    fig, ax = plt.subplots(figsize=(12, 8))
                    pos = nx.spring_layout(G, seed=42)
                    
                    # Nodi entry-level
                    entry_nodes = [n for n, d in G.nodes(data=True) if d.get('is_entry')]
                    nx.draw_networkx_nodes(G, pos, nodelist=entry_nodes, 
                                         node_color='lightgreen', node_size=700, alpha=0.8)
                    
                    # Nodo selezionato
                    selected_nodes = [n for n, d in G.nodes(data=True) if d.get('is_selected') and not d.get('is_entry')]
                    if selected_nodes:
                        nx.draw_networkx_nodes(G, pos, nodelist=selected_nodes, 
                                             node_color='yellow', node_size=700, alpha=0.8)
                    
                    # Altri nodi
                    other_nodes = [n for n, d in G.nodes(data=True) 
                                  if not d.get('is_entry') and not d.get('is_selected')]
                    nx.draw_networkx_nodes(G, pos, nodelist=other_nodes, 
                                         node_color='skyblue', node_size=700, alpha=0.8)
                    
                    # Etichette
                    labels = {n: f"{n}\n({d['count']} camere)\n€{d['adr']:.0f}" 
                             for n, d in G.nodes(data=True)}
                    nx.draw_networkx_labels(G, pos, labels=labels, font_size=9)
                    
                    # Archi
                    nx.draw_networkx_edges(G, pos, arrows=True, arrowsize=20, 
                                         width=2, alpha=0.7, edge_color='gray')
                    
                    plt.title(f"Gerarchia di Upgrade - {selected_hotel_name}")
                    plt.axis('off')
                    st.pyplot(fig)
                except Exception as e:
                    st.error(f"Errore nella visualizzazione grafico: {e}")
                    st.warning("Visualizzazione grafico non disponibile. Assicurati di avere le relazioni di upgrade configurate correttamente.")
            else:
                st.warning(f"Nessuna tipologia configurata per {selected_hotel_name}")
        else:
            st.warning("Seleziona un hotel configurato per visualizzare la panoramica completa.")

# Footer con informazioni
st.sidebar.markdown("---")
st.sidebar.info("""
### Hotel Upgrade Advisor Pro v2.0
Sviluppato per ottimizzare le decisioni di revenue management alberghiero.

**Caratteristiche principali:**
- Integrazione con Business Intelligence
- Riconoscimento automatico delle tipologie
- Controllo dell'inventario e overbooking
- Gestione tipologie entry-level
- Dashboard completa per il revenue management

**© 2025 Hotel Solutions**
""")
